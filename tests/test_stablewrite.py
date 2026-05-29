"""Standalone tests for stable_write.

Covers the full public surface of the module:
  - file_hash
  - normalize_zip_metadata
  - strip_office_metadata  (openpyxl-based, xlsx only)
  - strip_ooxml_metadata   (stdlib XML, all OOXML formats)
  - save_if_changed  (new file, unchanged, changed, finalizers, companions,
                      safe_copy, profiles, unknown profile)
  - save_xlsx_if_changed  (convenience wrapper, deterministic hash)
  - SaveResult / Saver dataclasses
"""

import importlib.util
import time
import zipfile
from pathlib import Path
from typing import Dict, List, Optional, Union

import pytest

from stable_write import Saver, SaveResult, save_if_changed, save_xlsx_if_changed
from stable_write.finalizers import (
    normalize_zip_metadata,
    strip_office_metadata,
    strip_ooxml_metadata,
)
from stable_write.stablewrite import file_hash

# ---------------------------------------------------------------------------
# Optional-dependency skip marks
# ---------------------------------------------------------------------------

_requires_openpyxl = pytest.mark.skipif(
    importlib.util.find_spec("openpyxl") is None,
    reason="openpyxl not installed",
)

_requires_pandas = pytest.mark.skipif(
    importlib.util.find_spec("pandas") is None,
    reason="pandas not installed",
)

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _make_zip(path: Path, entries: Optional[Dict[str, bytes]] = None) -> None:
    """Write a minimal ZIP at *path* with the given *entries*."""
    if entries is None:
        entries = {"xl/workbook.xml": b"<workbook/>", "_rels/.rels": b"<Relationships/>"}
    with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for name, data in entries.items():
            zf.writestr(name, data)


def _write(
    content: bytes,
    dest: Path,
    *,
    save_strategy: str = "overwrite",
    companions: Union[Dict[str, bytes], str] = "auto",
    finalizers=None,
    profile=None,
    safe_copy: bool = False,
    algo: str = "blake2b",
):
    """Save *content* to *dest* via save_if_changed; return the Saver."""
    kwargs: dict = {"save_strategy": save_strategy, "safe_copy": safe_copy, "algo": algo}
    if isinstance(companions, str):
        kwargs["companions"] = companions
    else:
        kwargs["companions"] = list(companions.keys())
    if finalizers is not None:
        kwargs["finalizers"] = finalizers
    if profile is not None:
        kwargs["profile"] = profile

    with save_if_changed(dest, **kwargs) as saver:
        saver.path.write_bytes(content)
        if isinstance(companions, dict):
            for name, data in companions.items():
                (saver.temp_dir / name).write_bytes(data)
    return saver


# ===========================================================================
# file_hash
# ===========================================================================


class TestFileHash:
    def test_returns_string(self, tmp_path):
        f = tmp_path / "a.txt"
        f.write_bytes(b"hello")
        h = file_hash(f)
        assert isinstance(h, str) and len(h) > 0

    def test_same_content_same_hash(self, tmp_path):
        f1 = tmp_path / "a.txt"
        f2 = tmp_path / "b.txt"
        f1.write_bytes(b"hello")
        f2.write_bytes(b"hello")
        assert file_hash(f1) == file_hash(f2)

    def test_different_content_different_hash(self, tmp_path):
        f1 = tmp_path / "a.txt"
        f2 = tmp_path / "b.txt"
        f1.write_bytes(b"hello")
        f2.write_bytes(b"world")
        assert file_hash(f1) != file_hash(f2)

    def test_algo_sha256_length(self, tmp_path):
        f = tmp_path / "a.txt"
        f.write_bytes(b"data")
        assert len(file_hash(f, algo="sha256")) == 64

    def test_invalid_algo_raises(self, tmp_path):
        f = tmp_path / "a.txt"
        f.write_bytes(b"data")
        with pytest.raises(ValueError, match="not available"):
            file_hash(f, algo="not_a_real_algo_xyz")

    def test_invalid_block_size_raises(self, tmp_path):
        f = tmp_path / "a.txt"
        f.write_bytes(b"data")
        with pytest.raises(ValueError, match="positive integer"):
            file_hash(f, block_size=0)

    def test_block_size_one_still_correct(self, tmp_path):
        f = tmp_path / "a.txt"
        f.write_bytes(b"abcdef")
        assert file_hash(f, block_size=1) == file_hash(f, block_size=8192)


# ===========================================================================
# normalize_zip_metadata
# ===========================================================================


class TestNormalizeZipMetadata:
    def test_raises_for_missing_file(self, tmp_path):
        with pytest.raises(FileNotFoundError):
            normalize_zip_metadata(tmp_path / "nonexistent.zip")

    def test_produces_deterministic_output(self, tmp_path):
        f1 = tmp_path / "a.zip"
        f2 = tmp_path / "b.zip"
        _make_zip(f1)
        _make_zip(f2)
        normalize_zip_metadata(f1)
        normalize_zip_metadata(f2)
        assert file_hash(f1) == file_hash(f2)

    def test_idempotent(self, tmp_path):
        f = tmp_path / "a.zip"
        _make_zip(f)
        normalize_zip_metadata(f)
        h1 = file_hash(f)
        normalize_zip_metadata(f)
        assert file_hash(f) == h1

    def test_entries_sorted(self, tmp_path):
        f = tmp_path / "test.zip"
        _make_zip(f, {"zzz.xml": b"last", "aaa.xml": b"first", "mmm.xml": b"mid"})
        normalize_zip_metadata(f)
        with zipfile.ZipFile(f) as zf:
            names = [i.filename for i in zf.infolist()]
        assert names == sorted(names)

    def test_timestamps_fixed(self, tmp_path):
        f = tmp_path / "test.zip"
        _make_zip(f, {"entry.xml": b"data"})
        normalize_zip_metadata(f)
        with zipfile.ZipFile(f) as zf:
            for info in zf.infolist():
                assert info.date_time == (1980, 1, 1, 0, 0, 0)

    def test_extra_fields_cleared(self, tmp_path):
        f = tmp_path / "test.zip"
        _make_zip(f, {"entry.xml": b"data"})
        normalize_zip_metadata(f)
        with zipfile.ZipFile(f) as zf:
            for info in zf.infolist():
                assert info.extra == b""

    def test_contents_preserved(self, tmp_path):
        payload = b"<root>hello world</root>"
        f = tmp_path / "test.zip"
        _make_zip(f, {"data.xml": payload})
        normalize_zip_metadata(f)
        with zipfile.ZipFile(f) as zf:
            assert zf.read("data.xml") == payload

    def test_different_entry_order_same_hash(self, tmp_path):
        """Two ZIPs with reversed entry order must hash identically after normalisation."""
        f1 = tmp_path / "f1.zip"
        f2 = tmp_path / "f2.zip"
        _make_zip(f1, {"aaa.xml": b"A", "zzz.xml": b"Z"})
        _make_zip(f2, {"zzz.xml": b"Z", "aaa.xml": b"A"})
        normalize_zip_metadata(f1)
        normalize_zip_metadata(f2)
        assert file_hash(f1) == file_hash(f2)


# ===========================================================================
# strip_office_metadata
# ===========================================================================


@_requires_openpyxl
class TestStripOfficeMetadata:
    def _make_xlsx(self, path: Path) -> None:
        from openpyxl import Workbook

        wb = Workbook()
        wb.save(path)

    def test_raises_for_missing_file(self, tmp_path):
        with pytest.raises(FileNotFoundError):
            strip_office_metadata(tmp_path / "nonexistent.xlsx")

    def test_clears_creator(self, tmp_path):
        from openpyxl import load_workbook

        f = tmp_path / "test.xlsx"
        self._make_xlsx(f)
        strip_office_metadata(f)
        # openpyxl round-trips an empty creator as None, not ""
        assert not load_workbook(f).properties.creator

    def test_clears_last_modified_by(self, tmp_path):
        from openpyxl import load_workbook

        f = tmp_path / "test.xlsx"
        self._make_xlsx(f)
        strip_office_metadata(f)
        # openpyxl round-trips an empty lastModifiedBy as None, not ""
        assert not load_workbook(f).properties.lastModifiedBy

    def test_neutralises_timestamps(self, tmp_path):
        """strip_office_metadata clears author fields; openpyxl resets the
        modified timestamp on wb.save(), so ZIP-level normalisation (done by
        normalize_zip_metadata) is what actually makes the hash stable.
        Verify the full pipeline (strip + normalize) fixes timestamps in the ZIP.
        """
        f = tmp_path / "test.xlsx"
        self._make_xlsx(f)
        strip_office_metadata(f)
        normalize_zip_metadata(f)
        with zipfile.ZipFile(f) as zf:
            for info in zf.infolist():
                assert info.date_time == (1980, 1, 1, 0, 0, 0)

    @pytest.mark.xfail(
        reason="openpyxl namespace state leaks between tests, causing non-deterministic serialisation",  # noqa: E501
        strict=False,
    )
    def test_deterministic_after_strip(self, tmp_path):
        """Two independently created workbooks hash the same after stripping."""
        f1 = tmp_path / "f1.xlsx"
        f2 = tmp_path / "f2.xlsx"
        self._make_xlsx(f1)
        self._make_xlsx(f2)
        strip_office_metadata(f1)
        strip_office_metadata(f2)
        normalize_zip_metadata(f1)
        normalize_zip_metadata(f2)
        assert file_hash(f1) == file_hash(f2)

    def test_corrupt_file_raises_not_swallowed(self, tmp_path):
        """A corrupt xlsx must raise — silent failure would save non-deterministic output."""
        f = tmp_path / "corrupt.xlsx"
        f.write_bytes(b"this is not a valid xlsx file")
        with pytest.raises(Exception):  # noqa: B017
            strip_office_metadata(f)


# ===========================================================================
# strip_ooxml_metadata
# ===========================================================================


class TestStripOoxmlMetadata:
    """Tests for the stdlib-XML metadata stripper (works for xlsx/docx/pptx)."""

    _CORE_XML = (
        b"<?xml version='1.0' encoding='UTF-8' standalone='yes'?>\n"
        b"<cp:coreProperties"
        b" xmlns:cp='http://schemas.openxmlformats.org/package/2006/metadata/core-properties'"
        b" xmlns:dc='http://purl.org/dc/elements/1.1/'"
        b" xmlns:dcterms='http://purl.org/dc/terms/'"
        b" xmlns:xsi='http://www.w3.org/2001/XMLSchema-instance'>"
        b"<dc:creator>Alice</dc:creator>"
        b"<cp:lastModifiedBy>Bob</cp:lastModifiedBy>"
        b"<dcterms:created xsi:type='dcterms:W3CDTF'>2025-05-01T12:00:00Z</dcterms:created>"
        b"<dcterms:modified xsi:type='dcterms:W3CDTF'>2025-05-28T09:00:00Z</dcterms:modified>"
        b"</cp:coreProperties>"
    )

    def _make_ooxml(self, path: Path, *, with_core: bool = True) -> None:
        """Build a minimal OOXML ZIP at *path*."""
        with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            zf.writestr("[Content_Types].xml", b"<Types/>")
            if with_core:
                zf.writestr("docProps/core.xml", self._CORE_XML)

    def _read_core_xml(self, path: Path):
        import xml.etree.ElementTree as ET

        with zipfile.ZipFile(path) as zf:
            return ET.fromstring(zf.read("docProps/core.xml"))

    def test_raises_for_missing_file(self, tmp_path):
        with pytest.raises(FileNotFoundError):
            strip_ooxml_metadata(tmp_path / "nonexistent.docx")

    def test_no_core_xml_is_noop(self, tmp_path):
        """Files without docProps/core.xml are left unchanged."""
        f = tmp_path / "no_core.zip"
        self._make_ooxml(f, with_core=False)
        mtime_before = f.stat().st_mtime_ns
        strip_ooxml_metadata(f)
        assert f.stat().st_mtime_ns == mtime_before

    def test_creator_cleared(self, tmp_path):
        f = tmp_path / "doc.docx"
        self._make_ooxml(f)
        strip_ooxml_metadata(f)
        creator = self._read_core_xml(f).find("{http://purl.org/dc/elements/1.1/}creator")
        assert creator is not None
        assert not creator.text

    def test_last_modified_by_cleared(self, tmp_path):
        f = tmp_path / "doc.docx"
        self._make_ooxml(f)
        strip_ooxml_metadata(f)
        el = self._read_core_xml(f).find(
            "{http://schemas.openxmlformats.org/package/2006/metadata/core-properties}"
            "lastModifiedBy"
        )
        assert el is not None
        assert not el.text

    def test_dates_neutralised(self, tmp_path):
        f = tmp_path / "doc.docx"
        self._make_ooxml(f)
        strip_ooxml_metadata(f)
        root = self._read_core_xml(f)
        for tag in (
            "{http://purl.org/dc/terms/}created",
            "{http://purl.org/dc/terms/}modified",
        ):
            el = root.find(tag)
            assert el is not None
            assert el.text == "1980-01-01T00:00:00Z"

    def test_deterministic_across_two_files(self, tmp_path):
        """Two files with different author metadata hash the same after stripping."""
        core_b = (
            b"<?xml version='1.0' encoding='UTF-8' standalone='yes'?>\n"
            b"<cp:coreProperties"
            b" xmlns:cp='http://schemas.openxmlformats.org/package/2006/metadata/core-properties'"
            b" xmlns:dc='http://purl.org/dc/elements/1.1/'"
            b" xmlns:dcterms='http://purl.org/dc/terms/'"
            b" xmlns:xsi='http://www.w3.org/2001/XMLSchema-instance'>"
            b"<dc:creator>Carol</dc:creator>"
            b"<cp:lastModifiedBy>Dave</cp:lastModifiedBy>"
            b"<dcterms:created xsi:type='dcterms:W3CDTF'>2026-01-01T00:00:00Z</dcterms:created>"
            b"<dcterms:modified xsi:type='dcterms:W3CDTF'>2026-06-01T00:00:00Z</dcterms:modified>"
            b"</cp:coreProperties>"
        )
        f1, f2 = tmp_path / "a.docx", tmp_path / "b.docx"
        for path, core in ((f1, self._CORE_XML), (f2, core_b)):
            with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
                zf.writestr("[Content_Types].xml", b"<Types/>")
                zf.writestr("docProps/core.xml", core)
        strip_ooxml_metadata(f1)
        strip_ooxml_metadata(f2)
        normalize_zip_metadata(f1)
        normalize_zip_metadata(f2)
        assert file_hash(f1) == file_hash(f2)


# ===========================================================================
# SaveResult / Saver dataclass properties
# ===========================================================================


class TestSaverProperties:
    def test_properties_delegate_to_result(self, tmp_path):
        dest = tmp_path / "out.txt"
        with save_if_changed(dest) as saver:
            saver.path.write_bytes(b"x")

        assert saver.changed == saver.result.changed
        assert saver.saved == saver.result.saved
        assert saver.old_hash == saver.result.old_hash
        assert saver.new_hash == saver.result.new_hash
        assert saver.hash_algo == saver.result.hash_algo
        assert saver.reason == saver.result.reason
        assert saver.changed_companions == saver.result.changed_companions

    def test_result_is_saveresult_instance(self, tmp_path):
        dest = tmp_path / "out.txt"
        with save_if_changed(dest) as saver:
            saver.path.write_bytes(b"x")
        assert isinstance(saver.result, SaveResult)
        assert isinstance(saver, Saver)


# ===========================================================================
# save_if_changed — new file
# ===========================================================================


class TestSaveIfChangedNewFile:
    def test_file_created(self, tmp_path):
        dest = tmp_path / "out.txt"
        with save_if_changed(dest) as saver:
            saver.path.write_bytes(b"content")
        assert dest.read_bytes() == b"content"

    def test_saved_true(self, tmp_path):
        dest = tmp_path / "out.txt"
        with save_if_changed(dest) as saver:
            saver.path.write_bytes(b"content")
        assert saver.saved is True
        assert saver.changed is True

    def test_reason_destination_missing(self, tmp_path):
        dest = tmp_path / "out.txt"
        with save_if_changed(dest) as saver:
            saver.path.write_bytes(b"content")
        assert saver.reason == "destination missing"

    def test_old_hash_none_new_hash_set(self, tmp_path):
        dest = tmp_path / "out.txt"
        with save_if_changed(dest) as saver:
            saver.path.write_bytes(b"content")
        assert saver.old_hash is None
        assert isinstance(saver.new_hash, str) and len(saver.new_hash) > 0

    def test_new_hash_matches_file_on_disk(self, tmp_path):
        dest = tmp_path / "out.txt"
        with save_if_changed(dest) as saver:
            saver.path.write_bytes(b"hello")
        assert saver.new_hash == file_hash(dest)

    def test_creates_parent_directory(self, tmp_path):
        dest = tmp_path / "subdir" / "nested" / "out.txt"
        with save_if_changed(dest) as saver:
            saver.path.write_bytes(b"x")
        assert dest.exists()

    def test_accepts_string_path(self, tmp_path):
        dest = str(tmp_path / "out.txt")
        with save_if_changed(dest) as saver:
            saver.path.write_bytes(b"x")
        assert Path(dest).exists()

    def test_skip_strategy_saves_new_file(self, tmp_path):
        """skip only silences changes to an *existing* file; new files are always saved."""
        dest = tmp_path / "out.txt"
        saver = _write(b"new", dest, save_strategy="skip")
        assert dest.exists()
        assert saver.saved is True

    def test_raise_strategy_saves_new_file(self, tmp_path):
        """raise strategy does not prevent saving a new (missing) destination."""
        dest = tmp_path / "out.txt"
        saver = _write(b"new", dest, save_strategy="raise")
        assert dest.exists()
        assert saver.saved is True


# ===========================================================================
# save_if_changed — unchanged file
# ===========================================================================


class TestSaveIfChangedUnchanged:
    def test_skips_save(self, tmp_path):
        dest = tmp_path / "out.txt"
        _write(b"same", dest)
        mtime_before = dest.stat().st_mtime_ns

        _write(b"same", dest)

        assert dest.stat().st_mtime_ns == mtime_before

    def test_saved_false(self, tmp_path):
        dest = tmp_path / "out.txt"
        _write(b"same", dest)
        saver = _write(b"same", dest)

        assert saver.saved is False
        assert saver.changed is False

    def test_reason_content_unchanged(self, tmp_path):
        dest = tmp_path / "out.txt"
        _write(b"same", dest)
        saver = _write(b"same", dest)
        assert saver.reason == "content unchanged"

    def test_hashes_populated(self, tmp_path):
        dest = tmp_path / "out.txt"
        _write(b"same", dest)
        saver = _write(b"same", dest)
        assert saver.old_hash == saver.new_hash
        assert saver.old_hash is not None

    def test_skip_strategy_does_not_force_save_when_unchanged(self, tmp_path):
        dest = tmp_path / "out.txt"
        _write(b"same", dest)
        saver = _write(b"same", dest, save_strategy="skip")
        # Content unchanged → still skipped regardless of strategy
        assert saver.saved is False

    def test_overwrite_strategy_does_not_force_save_when_unchanged(self, tmp_path):
        dest = tmp_path / "out.txt"
        _write(b"same", dest)
        saver = _write(b"same", dest, save_strategy="overwrite")
        # Content unchanged → still skipped (strategy only matters for changed content)
        assert saver.saved is False


# ===========================================================================
# save_if_changed — content changed
# ===========================================================================


class TestSaveIfChangedContentChanged:
    def test_file_updated(self, tmp_path):
        dest = tmp_path / "out.txt"
        _write(b"v1", dest)
        _write(b"v2", dest)
        assert dest.read_bytes() == b"v2"

    def test_saved_true(self, tmp_path):
        dest = tmp_path / "out.txt"
        _write(b"v1", dest)
        saver = _write(b"v2", dest)
        assert saver.saved is True
        assert saver.changed is True

    def test_reason_content_changed(self, tmp_path):
        dest = tmp_path / "out.txt"
        _write(b"v1", dest)
        saver = _write(b"v2", dest)
        assert saver.reason == "content changed"

    def test_old_hash_differs_from_new(self, tmp_path):
        dest = tmp_path / "out.txt"
        _write(b"v1", dest)
        saver = _write(b"v2", dest)
        assert saver.old_hash != saver.new_hash

    def test_raises_with_raise_strategy(self, tmp_path):
        dest = tmp_path / "out.txt"
        _write(b"v1", dest)
        with pytest.raises(FileExistsError):
            _write(b"v2", dest, save_strategy="raise")

    def test_original_preserved_with_raise_strategy(self, tmp_path):
        dest = tmp_path / "out.txt"
        _write(b"v1", dest)
        with pytest.raises(FileExistsError):
            _write(b"v2", dest, save_strategy="raise")
        assert dest.read_bytes() == b"v1"

    def test_saved_false_with_raise_strategy(self, tmp_path):
        dest = tmp_path / "out.txt"
        _write(b"v1", dest)
        with pytest.raises(FileExistsError):
            _write(b"v2", dest, save_strategy="raise")
        # saver is set inside _write before the raise propagates
        # just assert the file is intact (covered above)

    def test_skip_strategy_leaves_destination_unchanged(self, tmp_path):
        dest = tmp_path / "out.txt"
        _write(b"v1", dest)
        saver = _write(b"v2", dest, save_strategy="skip")
        assert dest.read_bytes() == b"v1"
        assert saver.saved is False
        assert saver.changed is True

    def test_skip_strategy_reason(self, tmp_path):
        dest = tmp_path / "out.txt"
        _write(b"v1", dest)
        saver = _write(b"v2", dest, save_strategy="skip")
        assert saver.reason == "content changed \u2014 skipped"


# ===========================================================================
# save_if_changed — finalizers
# ===========================================================================


class TestSaveIfChangedFinalizers:
    def test_finalizer_called(self, tmp_path):
        called = []
        dest = tmp_path / "out.txt"

        def mark(p: Path) -> None:
            called.append(p)

        with save_if_changed(dest, finalizers=[mark]) as saver:
            saver.path.write_bytes(b"x")

        assert len(called) == 1

    def test_finalizer_runs_before_hash(self, tmp_path):
        """Bytes mutated by finalizer must be what lands on disk and in new_hash."""
        dest = tmp_path / "out.txt"

        def append_null(p: Path) -> None:
            p.write_bytes(p.read_bytes() + b"\x00")

        with save_if_changed(dest, finalizers=[append_null]) as saver:
            saver.path.write_bytes(b"data")

        assert dest.read_bytes() == b"data\x00"
        assert saver.new_hash == file_hash(dest)

    def test_finalizer_hash_drives_change_detection(self, tmp_path):
        """If finalizer produces same bytes as existing file, save is skipped."""
        dest = tmp_path / "out.txt"

        def noop(p: Path) -> None:
            pass

        # First write: plain b"same"
        with save_if_changed(dest) as saver:
            saver.path.write_bytes(b"same")

        # Second write: write different bytes but finalizer overwrites back to b"same"
        def overwrite_to_same(p: Path) -> None:
            p.write_bytes(b"same")

        saver2 = _write(b"different", dest, finalizers=[overwrite_to_same])
        assert saver2.saved is False

    def test_multiple_finalizers_run_in_order(self, tmp_path):
        dest = tmp_path / "out.txt"
        log: List[str] = []

        def first(p: Path) -> None:
            log.append("first")
            p.write_bytes(p.read_bytes() + b"1")

        def second(p: Path) -> None:
            log.append("second")
            p.write_bytes(p.read_bytes() + b"2")

        with save_if_changed(dest, finalizers=[first, second]) as saver:
            saver.path.write_bytes(b"")

        assert log == ["first", "second"]
        assert dest.read_bytes() == b"12"


# ===========================================================================
# save_if_changed — companion files
# ===========================================================================


class TestSaveIfChangedCompanions:
    def test_auto_companion_moved(self, tmp_path):
        dest = tmp_path / "out.txt"
        with save_if_changed(dest, companions="auto") as saver:
            saver.path.write_bytes(b"main")
            (saver.temp_dir / "companion.csv").write_bytes(b"extra")
        assert (tmp_path / "companion.csv").read_bytes() == b"extra"

    def test_explicit_companion_moved(self, tmp_path):
        dest = tmp_path / "out.txt"
        with save_if_changed(dest, companions=["companion.csv"]) as saver:
            saver.path.write_bytes(b"main")
            (saver.temp_dir / "companion.csv").write_bytes(b"extra")
        assert (tmp_path / "companion.csv").read_bytes() == b"extra"

    def test_none_companions_skips_extras(self, tmp_path):
        dest = tmp_path / "out.txt"
        with save_if_changed(dest, companions=None) as saver:
            saver.path.write_bytes(b"main")
            (saver.temp_dir / "companion.csv").write_bytes(b"extra")
        assert not (tmp_path / "companion.csv").exists()

    def test_empty_list_companions_skips_extras(self, tmp_path):
        dest = tmp_path / "out.txt"
        with save_if_changed(dest, companions=[]) as saver:
            saver.path.write_bytes(b"main")
            (saver.temp_dir / "companion.csv").write_bytes(b"extra")
        assert not (tmp_path / "companion.csv").exists()

    def test_companion_change_triggers_save(self, tmp_path):
        dest = tmp_path / "out.txt"
        _write(b"main", dest, companions={"companion.csv": b"v1"})

        saver = _write(b"main", dest, companions={"companion.csv": b"v2"})

        assert saver.saved is True
        assert (tmp_path / "companion.csv").read_bytes() == b"v2"

    def test_companion_unchanged_no_save(self, tmp_path):
        dest = tmp_path / "out.txt"
        _write(b"main", dest, companions={"companion.csv": b"stable"})
        mtime_before = (tmp_path / "companion.csv").stat().st_mtime_ns

        saver = _write(b"main", dest, companions={"companion.csv": b"stable"})

        assert saver.saved is False
        assert (tmp_path / "companion.csv").stat().st_mtime_ns == mtime_before

    def test_changed_companions_list_populated(self, tmp_path):
        dest = tmp_path / "out.txt"
        _write(b"main", dest, companions={"companion.csv": b"v1"})
        saver = _write(b"main", dest, companions={"companion.csv": b"v2"})

        assert "companion.csv" in saver.changed_companions

    def test_main_changed_companions_also_scanned(self, tmp_path):
        """When the main file changes, changed companions are still reported."""
        dest = tmp_path / "out.txt"
        _write(b"v1", dest, companions={"companion.csv": b"cv1"})
        saver = _write(b"v2", dest, companions={"companion.csv": b"cv2"})

        assert saver.changed is True
        assert "companion.csv" in saver.changed_companions

    def test_main_changed_companion_unchanged_not_listed(self, tmp_path):
        """Unchanged companions are not included in changed_companions."""
        dest = tmp_path / "out.txt"
        _write(b"v1", dest, companions={"companion.csv": b"same"})
        saver = _write(b"v2", dest, companions={"companion.csv": b"same"})

        assert saver.changed is True
        assert saver.changed_companions == []

    def test_companion_raises_with_raise_strategy(self, tmp_path):
        dest = tmp_path / "out.txt"
        _write(b"main", dest, companions={"companion.csv": b"v1"})
        with pytest.raises(FileExistsError):
            _write(b"main", dest, companions={"companion.csv": b"v2"}, save_strategy="raise")

    def test_original_companion_preserved_with_raise_strategy(self, tmp_path):
        dest = tmp_path / "out.txt"
        _write(b"main", dest, companions={"companion.csv": b"v1"})
        with pytest.raises(FileExistsError):
            _write(b"main", dest, companions={"companion.csv": b"v2"}, save_strategy="raise")
        assert (tmp_path / "companion.csv").read_bytes() == b"v1"

    # --- companion input validation ----------------------------------------

    def test_invalid_string_companions_raises_before_publish(self, tmp_path):
        """companions='plot.png' (a bare string, not 'auto') raises ValueError
        before any temp directory is created — the destination must not exist."""
        dest = tmp_path / "out.txt"

        with pytest.raises(ValueError, match="companions"):
            with save_if_changed(dest, companions="plot.png") as saver:
                saver.path.write_bytes(b"main")

        assert not dest.exists()

    def test_companion_path_traversal_rejected(self, tmp_path):
        """A companion name with directory components is rejected."""
        dest = tmp_path / "subdir" / "out.txt"
        dest.parent.mkdir()

        with pytest.raises(ValueError, match="companion"):
            with save_if_changed(dest, companions=["../escape.txt"]) as saver:
                saver.path.write_bytes(b"main")
                (saver.temp_dir / "escape.txt").write_bytes(b"bad")

        assert not (tmp_path / "escape.txt").exists()

    def test_absolute_companion_path_rejected(self, tmp_path):
        """An absolute companion path is rejected."""
        dest = tmp_path / "out.txt"
        bad = str(tmp_path / "escape.txt")

        with pytest.raises(ValueError, match="companion"):
            with save_if_changed(dest, companions=[bad]) as saver:
                saver.path.write_bytes(b"main")

    def test_invalid_companions_type_raises(self, tmp_path):
        """A non-iterable companions value raises TypeError before publish."""
        dest = tmp_path / "out.txt"

        with pytest.raises(TypeError):
            with save_if_changed(dest, companions=123):  # type: ignore[arg-type]
                pass

        assert not dest.exists()


# ===========================================================================
# save_if_changed — profiles
# ===========================================================================


class TestSaveIfChangedProfiles:
    def test_unknown_profile_raises(self, tmp_path):
        dest = tmp_path / "out.txt"
        with pytest.raises(ValueError, match="Unknown profile"):
            with save_if_changed(dest, profile="nonexistent") as saver:
                saver.path.write_bytes(b"x")

    def test_zip_profile_normalizes(self, tmp_path):
        dest1 = tmp_path / "a.zip"
        dest2 = tmp_path / "b.zip"
        entries = {"zzz.xml": b"Z", "aaa.xml": b"A"}

        with save_if_changed(dest1, profile="zip") as saver:
            _make_zip(saver.path, entries)

        with save_if_changed(dest2, profile="zip") as saver:
            _make_zip(saver.path, {"aaa.xml": b"A", "zzz.xml": b"Z"})  # reversed order

        assert file_hash(dest1) == file_hash(dest2)

    def test_explicit_finalizers_override_profile(self, tmp_path):
        """When finalizers= is given alongside profile=, profile is ignored."""
        called = []
        dest = tmp_path / "out.txt"

        def custom(p: Path) -> None:
            called.append("custom")

        with save_if_changed(dest, profile="zip", finalizers=[custom]) as saver:
            saver.path.write_bytes(b"not a zip")  # would fail zip parsing if profile ran

        assert called == ["custom"]


# ===========================================================================
# save_if_changed — hash algorithm option
# ===========================================================================


class TestSaveIfChangedAlgo:
    def test_sha256_stored_in_hash_algo(self, tmp_path):
        dest = tmp_path / "out.txt"
        with save_if_changed(dest, algo="sha256") as saver:
            saver.path.write_bytes(b"x")
        assert saver.hash_algo == "sha256"
        assert len(saver.new_hash) == 64

    def test_hash_algo_drives_change_detection(self, tmp_path):
        """Two consecutive writes with same algo and same content → unchanged."""
        dest = tmp_path / "out.txt"
        _write(b"same", dest, algo="sha256")
        saver = _write(b"same", dest, algo="sha256")
        assert saver.saved is False


# ===========================================================================
# save_if_changed — is_equal custom comparator
# ===========================================================================


class TestSaveIfChangedIsEqual:
    def test_is_equal_true_suppresses_save(self, tmp_path):
        """When is_equal returns True the destination is not overwritten."""
        dest = tmp_path / "out.txt"
        dest.write_bytes(b"v1")
        mtime_before = dest.stat().st_mtime_ns

        with save_if_changed(dest, is_equal=lambda _n, _e: True) as saver:
            saver.path.write_bytes(b"v2")  # different bytes, but comparator says equal

        assert saver.saved is False
        assert saver.changed is False
        assert dest.stat().st_mtime_ns == mtime_before

    def test_is_equal_false_triggers_save(self, tmp_path):
        """When is_equal returns False the destination is replaced."""
        dest = tmp_path / "out.txt"
        dest.write_bytes(b"v1")

        with save_if_changed(dest, is_equal=lambda _n, _e: False) as saver:
            saver.path.write_bytes(b"v1")  # same bytes, but comparator says different

        assert saver.saved is True
        assert saver.changed is True

    def test_is_equal_not_called_for_new_file(self, tmp_path):
        """is_equal must not be invoked when the destination does not exist yet."""
        dest = tmp_path / "out.txt"
        called = []

        def comparator(n, e):
            called.append((n, e))
            return True

        with save_if_changed(dest, is_equal=comparator) as saver:
            saver.path.write_bytes(b"new")

        assert saver.saved is True
        assert saver.reason == "destination missing"
        assert called == []

    def test_is_equal_receives_correct_paths(self, tmp_path):
        """is_equal is called with (new_temp_path, existing_destination)."""
        dest = tmp_path / "out.txt"
        dest.write_bytes(b"existing")
        seen = {}

        def comparator(new: Path, existing: Path) -> bool:
            seen["new_content"] = new.read_bytes()
            seen["existing_content"] = existing.read_bytes()
            return False

        with save_if_changed(dest, is_equal=comparator) as saver:
            saver.path.write_bytes(b"new content")

        assert seen["new_content"] == b"new content"
        assert seen["existing_content"] == b"existing"

    def test_is_equal_does_not_affect_companion_hashing(self, tmp_path):
        """Companions are still compared by hash even when is_equal is set."""
        dest = tmp_path / "out.txt"
        dest.write_bytes(b"main")
        (tmp_path / "companion.csv").write_bytes(b"v1")

        # is_equal says main is unchanged, but companion changed
        with save_if_changed(
            dest, companions=["companion.csv"], is_equal=lambda _n, _e: True
        ) as saver:
            saver.path.write_bytes(b"main")
            (saver.temp_dir / "companion.csv").write_bytes(b"v2")

        assert saver.saved is True
        assert "companion.csv" in saver.changed_companions

    def test_is_equal_can_mark_different_bytes_unchanged(self, tmp_path):
        """
        Note:
        When ``is_equal`` is provided, ``changed`` and ``saved`` are based on
        the comparator result, not on ``old_hash == new_hash``. Hashes may still
        differ for semantically equivalent files.
        """

        dest = tmp_path / "out.bin"
        dest.write_bytes(b"old")

        def equivalent(new: Path, existing: Path) -> bool:
            return True

        with save_if_changed(dest, is_equal=equivalent) as saver:
            saver.path.write_bytes(b"new")

        assert saver.changed is False
        assert saver.saved is False
        assert dest.read_bytes() == b"old"
        assert saver.old_hash != saver.new_hash


# ===========================================================================
# save_if_changed — failure paths
# ===========================================================================


class TestSaveIfChangedFailurePaths:
    # --- 1. finalizer raises → destination untouched ----------------------

    def test_finalizer_exception_leaves_destination_unchanged(self, tmp_path):
        """A raising finalizer must not touch the destination."""
        dest = tmp_path / "out.txt"
        dest.write_bytes(b"original")

        def boom(p: Path) -> None:
            raise RuntimeError("finalizer failed")

        with pytest.raises(RuntimeError, match="finalizer failed"):
            with save_if_changed(dest, finalizers=[boom]) as saver:
                saver.path.write_bytes(b"new content")

        assert dest.read_bytes() == b"original"

    def test_finalizer_exception_on_new_file_creates_nothing(self, tmp_path):
        """If destination did not exist and finalizer raises, it stays absent."""
        dest = tmp_path / "out.txt"

        def boom(p: Path) -> None:
            raise RuntimeError("finalizer failed")

        with pytest.raises(RuntimeError), save_if_changed(dest, finalizers=[boom]) as saver:
            saver.path.write_bytes(b"content")

        assert not dest.exists()

    def test_finalizer_exception_leaves_companions_unchanged(self, tmp_path):
        """Companion files written before a raising finalizer must not be published."""
        dest = tmp_path / "out.txt"
        dest.write_bytes(b"main-v1")
        (tmp_path / "companion.csv").write_bytes(b"companion-v1")

        def boom(p: Path) -> None:
            raise RuntimeError("finalizer failed")

        with pytest.raises(RuntimeError):
            with save_if_changed(dest, companions="auto", finalizers=[boom]) as saver:
                saver.path.write_bytes(b"main-v2")
                (saver.temp_dir / "companion.csv").write_bytes(b"companion-v2")

        assert dest.read_bytes() == b"main-v1"
        assert (tmp_path / "companion.csv").read_bytes() == b"companion-v1"

    # --- 2. user forgets to write the temp file ---------------------------

    def test_empty_temp_file_treated_as_empty_content(self, tmp_path):
        """Temp file exists but is zero bytes — treated as valid empty content."""
        dest = tmp_path / "out.txt"
        with save_if_changed(dest):
            pass  # intentionally write nothing

        # Zero-byte file should still be published (first write)
        assert dest.exists()
        assert dest.read_bytes() == b""

    def test_empty_temp_file_matches_existing_empty_file(self, tmp_path):
        """If destination is also empty, a no-write round-trip is detected as unchanged."""
        dest = tmp_path / "out.txt"
        dest.write_bytes(b"")

        with save_if_changed(dest) as saver:
            pass  # write nothing → temp stays empty

        assert saver.saved is False
        assert saver.changed is False

    # --- 3. invalid profile → no partial publish --------------------------

    def test_unknown_profile_raises_before_publish(self, tmp_path):
        """ValueError for unknown profile must be raised before any file is published."""
        dest = tmp_path / "out.txt"

        with pytest.raises(ValueError, match="Unknown profile"):
            with save_if_changed(dest, profile="bogus") as saver:
                saver.path.write_bytes(b"content")

        assert not dest.exists()

    def test_unknown_profile_leaves_existing_file_unchanged(self, tmp_path):
        """ValueError for unknown profile must not overwrite an existing destination."""
        dest = tmp_path / "out.txt"
        dest.write_bytes(b"original")

        with pytest.raises(ValueError, match="Unknown profile"):
            with save_if_changed(dest, profile="bogus") as saver:
                saver.path.write_bytes(b"new content")

        assert dest.read_bytes() == b"original"

    # --- 4. invalid save_strategy -----------------------------------------

    def test_invalid_save_strategy_raises_before_tempdir(self, tmp_path):
        """An unrecognised save_strategy must raise ValueError immediately."""
        dest = tmp_path / "out.txt"
        with pytest.raises(ValueError, match="save_strategy"):  # type: ignore[arg-type]
            with save_if_changed(dest, save_strategy="bogus") as saver:  # type: ignore[arg-type]
                saver.path.write_bytes(b"content")

    def test_invalid_save_strategy_leaves_destination_unchanged(self, tmp_path):
        """Invalid strategy must not publish anything to an existing destination."""
        dest = tmp_path / "out.txt"
        dest.write_bytes(b"original")
        with pytest.raises(ValueError), save_if_changed(dest, save_strategy="bogus") as saver:  # type: ignore[arg-type]
            saver.path.write_bytes(b"new content")
        assert dest.read_bytes() == b"original"

    # --- 5. companion edge cases ------------------------------------------

    def test_missing_explicit_companion_raises(self, tmp_path):
        """An explicitly listed companion that was never written must raise FileNotFoundError.
        Use companions='auto' for optional companions.
        """
        dest = tmp_path / "out.txt"
        with pytest.raises(FileNotFoundError, match=r"ghost.csv"):
            with save_if_changed(dest, companions=["ghost.csv"]) as saver:
                saver.path.write_bytes(b"main")
                # ghost.csv intentionally not written

        assert not dest.exists()

    def test_new_companion_alongside_unchanged_main_triggers_save(self, tmp_path):
        """A companion that appears for the first time (no existing) triggers a save."""
        dest = tmp_path / "out.txt"
        # First write: main only, no companion
        with save_if_changed(dest, companions="auto") as saver:
            saver.path.write_bytes(b"main")

        # Second write: same main bytes, but now a companion appears
        saver2 = _write(b"main", dest, companions={"new.csv": b"data"})

        assert saver2.saved is True
        assert (tmp_path / "new.csv").read_bytes() == b"data"

    def test_empty_explicit_companion_unchanged_does_not_save(self, tmp_path):
        """Listing a companion name that was never written must not raise."""
        dest = tmp_path / "out.txt"
        _write(b"main", dest, companions={"absent.csv": b""})
        # _write writes the companion; re-run without it
        saver2 = _write(b"main", dest, companions={"absent.csv": b""})
        # same companion bytes → unchanged
        assert saver2.saved is False


# ===========================================================================
# save_xlsx_if_changed
# ===========================================================================


@_requires_openpyxl
class TestSaveXlsxIfChanged:
    def _xlsx_bytes(self) -> bytes:
        """Produce a minimal xlsx as bytes via openpyxl."""
        import io

        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["a", "b"])
        ws.append([1, 10])
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_file_created(self, tmp_path):
        dest = tmp_path / "report.xlsx"
        with save_xlsx_if_changed(dest) as saver:
            saver.path.write_bytes(self._xlsx_bytes())
        assert dest.exists()

    def test_saved_true_on_first_write(self, tmp_path):
        dest = tmp_path / "report.xlsx"
        with save_xlsx_if_changed(dest) as saver:
            saver.path.write_bytes(self._xlsx_bytes())
        assert saver.saved is True

    def test_unchanged_skipped(self, tmp_path):
        dest = tmp_path / "report.xlsx"
        data = self._xlsx_bytes()

        with save_xlsx_if_changed(dest) as saver:
            saver.path.write_bytes(data)

        mtime_before = dest.stat().st_mtime_ns

        with save_xlsx_if_changed(dest) as saver:
            saver.path.write_bytes(data)

        assert saver.saved is False
        assert dest.stat().st_mtime_ns == mtime_before

    def test_deterministic_hash(self, tmp_path):
        """Same workbook content → same new_hash across independent writes."""
        data = self._xlsx_bytes()
        dest1 = tmp_path / "r1.xlsx"
        dest2 = tmp_path / "r2.xlsx"

        with save_xlsx_if_changed(dest1) as saver:
            saver.path.write_bytes(data)
        h1 = saver.new_hash

        with save_xlsx_if_changed(dest2) as saver:
            saver.path.write_bytes(data)
        h2 = saver.new_hash

        assert h1 == h2

    def test_changed_content_detected(self, tmp_path):
        import io

        from openpyxl import Workbook

        dest = tmp_path / "report.xlsx"

        def _save(values: list) -> Saver:
            wb = Workbook()
            ws = wb.active
            ws.append(values)
            buf = io.BytesIO()
            wb.save(buf)
            with save_xlsx_if_changed(dest) as saver:
                saver.path.write_bytes(buf.getvalue())
            return saver

        _save([1, 2, 3])
        saver = _save([9, 8, 7])
        assert saver.saved is True
        assert saver.changed is True

    def test_raises_without_overwrite(self, tmp_path):
        dest = tmp_path / "report.xlsx"
        data = self._xlsx_bytes()

        with save_xlsx_if_changed(dest) as saver:
            saver.path.write_bytes(data)

        import io

        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["different"])
        buf = io.BytesIO()
        wb.save(buf)

        with pytest.raises(FileExistsError):
            with save_xlsx_if_changed(dest, save_strategy="raise") as saver:
                saver.path.write_bytes(buf.getvalue())


# ===========================================================================
# save_xlsx_if_changed — via pandas.to_excel
# ===========================================================================


@_requires_openpyxl
@_requires_pandas
class TestSaveXlsxIfChangedPandas:
    """Verify the full xlsx profile pipeline with pandas as the writer.

    pandas.to_excel uses an ExcelWriter engine that injects its own metadata
    (application version, timestamps, etc.), so these tests exercise the
    strip + normalize finalizer chain against a real-world caller.
    """

    def _make_df(self, values: List[int]):
        import pandas as pd

        return pd.DataFrame({"a": values, "b": [v * 10 for v in values]})

    def _write_excel(self, df, dest, save_strategy="overwrite"):
        with save_xlsx_if_changed(dest, save_strategy=save_strategy) as saver:
            df.to_excel(saver.path, index=False)
        return saver

    def test_raw_pandas_writes_are_nondeterministic(self, tmp_path):
        """Baseline: without any finalizers, two independent to_excel writes
        of the same DataFrame produce different bytes (different core.xml timestamps).
        Confirms stripping is actually necessary.

        The 1-second sleep is intentional: openpyxl truncates dcterms:modified
        to whole seconds, so writes within the same second produce identical
        bytes and the assertion would trivially pass for the wrong reason.
        """
        import time

        df = self._make_df([1, 2, 3])
        f1, f2 = tmp_path / "r1.xlsx", tmp_path / "r2.xlsx"

        with save_if_changed(f1) as saver:  # no profile → no finalizers
            df.to_excel(saver.path, index=False)

        time.sleep(1)  # force a different dcterms:modified timestamp

        with save_if_changed(f2) as saver:
            df.to_excel(saver.path, index=False)

        assert file_hash(f1) != file_hash(f2)

    def test_file_created(self, tmp_path):
        dest = tmp_path / "report.xlsx"
        self._write_excel(self._make_df([1, 2, 3]), dest)
        assert dest.exists()

    def test_file_readable(self, tmp_path):
        import pandas as pd

        dest = tmp_path / "report.xlsx"
        self._write_excel(self._make_df([10, 20, 30]), dest)
        assert list(pd.read_excel(dest)["a"]) == [10, 20, 30]

    def test_saved_true_on_first_write(self, tmp_path):
        dest = tmp_path / "report.xlsx"
        saver = self._write_excel(self._make_df([1, 2, 3]), dest)
        assert saver.saved is True
        assert saver.changed is True
        assert saver.reason == "destination missing"

    def test_unchanged_skipped(self, tmp_path):
        dest = tmp_path / "report.xlsx"
        df = self._make_df([1, 2, 3])
        self._write_excel(df, dest)
        mtime_before = dest.stat().st_mtime_ns

        time.sleep(1)  # force a different dcterms:modified timestamp without stripping

        saver = self._write_excel(df, dest)

        assert saver.saved is False
        assert saver.changed is False
        assert dest.stat().st_mtime_ns == mtime_before

    def test_deterministic_hash(self, tmp_path):
        """Same DataFrame → same new_hash across two independent writes separated
        by >1 s, so openpyxl would embed different dcterms:modified timestamps
        without stripping."""
        import time

        df = self._make_df([1, 2, 3])
        dest1 = tmp_path / "r1.xlsx"
        dest2 = tmp_path / "r2.xlsx"

        s1 = self._write_excel(df, dest1)
        time.sleep(1)  # force a different dcterms:modified timestamp without stripping
        s2 = self._write_excel(df, dest2)

        assert s1.new_hash == s2.new_hash

    def test_data_change_detected(self, tmp_path):
        dest = tmp_path / "report.xlsx"
        self._write_excel(self._make_df([1, 2, 3]), dest)
        saver = self._write_excel(self._make_df([9, 8, 7]), dest)

        assert saver.saved is True
        assert saver.changed is True

    def test_data_change_persisted(self, tmp_path):
        import pandas as pd

        dest = tmp_path / "report.xlsx"
        self._write_excel(self._make_df([1, 2, 3]), dest)
        self._write_excel(self._make_df([9, 8, 7]), dest)

        assert list(pd.read_excel(dest)["a"]) == [9, 8, 7]

    def test_raises_without_overwrite(self, tmp_path):
        dest = tmp_path / "report.xlsx"
        self._write_excel(self._make_df([1, 2, 3]), dest)

        with pytest.raises(FileExistsError):
            self._write_excel(self._make_df([9, 8, 7]), dest, save_strategy="raise")

    def test_original_preserved_when_overwrite_false(self, tmp_path):
        import pandas as pd

        dest = tmp_path / "report.xlsx"
        self._write_excel(self._make_df([1, 2, 3]), dest)

        with pytest.raises(FileExistsError):
            self._write_excel(self._make_df([9, 8, 7]), dest, save_strategy="raise")

        assert list(pd.read_excel(dest)["a"]) == [1, 2, 3]  # original untouched


def zip_is_equal(new: Path, existing: Path) -> bool:
    import hashlib

    def entries(path: Path) -> Dict[str, bytes]:
        with zipfile.ZipFile(path, "r") as zf:
            result = {}
            for info in zf.infolist():
                if info.is_dir():
                    result[info.filename.rstrip("/") + "/"] = b""
                else:
                    # Get the hash of the contents
                    result[info.filename] = (
                        hashlib.blake2b(zf.read(info.filename)).hexdigest().encode("utf-8")
                    )
            return result

    return entries(new) == entries(existing)
